def generatedata(data1, data2):
    from rich import pretty
    from rich import print
    pretty.install()

    # declare variables and import modules and packages
    global sheet_ranges
    data1name = ""
    data2name = ""
    from modules import vars
    import openpyxl
    wb = openpyxl.load_workbook(filename='data/large-dataset.xlsx')
    dataquantity = int(input('Enter the number of data points to generate: '))
    offset = (183 // dataquantity) - 1
    data1 = []
    data2 = []
    data1name = ""
    data2name = ""
    date1 = []
    date2 = []

    # data array logic
    for i in range(0, 2):
        print(i)
        if i == 0:
            print('Generating data set 1...')
        elif i == 1:
            print('Generating data set 2...')
        else:
            print('Maximum data sets reached')
            print('Clearing data sets...')
            print('Generating data set 1...')
            i = 0
            data1.clear()
            data2.clear()
            return

        # user chooses place to get data from
        print('Choose a location from the options below: ')
        print('1. Cambourne, United Kingdom (1987)')
        print('2. Heathrow, United Kingdom (1987)')
        print('3. Hurn, United Kingdom (1987)')
        print('4. Leeming, United Kingdom (1987)')
        print('5. Leuchars, United Kingdom (1987)')
        print('6. Cambourne, United Kingdom (2015)')
        print('7. Heathrow, United Kingdom (2015)')
        print('8. Hurn, United Kingdom (2015)')
        print('9. Leeming, United Kingdom (2015)')
        print('10. Leuchars, United Kingdom (2015)')

        choice = int(input('Enter your choice: '))
        if choice == 1:
            sheet_ranges = wb[vars.CAMBOURNE1987]

            if i == 0:
                data1name = vars.CAMBOURNE1987
            elif i == 1:
                data2name = vars.CAMBOURNE1987


        elif choice == 2:
            sheet_ranges = wb[vars.HEATHROW1987]

            if i == 0:
                data1name = vars.HEATHROW1987
            elif i == 1:
                data2name = vars.HEATHROW1987


        elif choice == 3:
            sheet_ranges = wb[vars.HURN1987]

            if i == 0:
                data1name = vars.HURN1987
            elif i == 1:
                data2name = vars.HURN1987


        elif choice == 4:
            sheet_ranges = wb[vars.LEEMING1987]
            if i == 0:
                data1name = vars.LEEMING1987
            elif i == 1:
                data2name = vars.LEEMING1987

        elif choice == 5:
            sheet_ranges = wb[vars.LEUCHARS1987]

            if i == 0:
                data1name = vars.LEUCHARS1987
            elif i == 1:
                data2name = vars.LEUCHARS1987


        elif choice == 6:
            sheet_ranges = wb[vars.CAMBOURNE2015]

            if i == 0:
                data1name = vars.CAMBOURNE2015
            elif i == 1:
                data2name = vars.CAMBOURNE2015


        elif choice == 7:
            sheet_ranges = wb[vars.HEATHROW2015]

            if i == 0:
                data1name = vars.HEATHROW2015
            elif i == 1:
                data2name = vars.HEATHROW2015


        elif choice == 8:
            sheet_ranges = wb[vars.HURN2015]

            if i == 0:
                data1name = vars.HURN2015
            elif i == 1:
                data2name = vars.HURN2015


        elif choice == 9:
            sheet_ranges = wb[vars.LEEMING2015]

            if i == 0:
                data1name = vars.LEEMING2015
            elif i == 1:
                data2name = vars.LEEMING2015


        elif choice == 10:
            sheet_ranges = wb[vars.LEUCHARS2015]

            if i == 0:
                data1name = vars.LEUCHARS2015
            elif i == 1:
                data2name = vars.LEUCHARS2015

        # string formatting
        if i == 0:
            data1name = data1name + ' -'
        elif i == 1:
            data2name = data2name + ' -'

        # user chooses what data to get
        print('Choose a parameter from the options below: ')
        print('1. Mean temperature')
        print('2. Total rainfall')
        print('3. Total sunshine')
        print('4. Mean wind speed')
        print('5. Relative humidity')
        print('6. Mean cloud cover')
        print('7. Mean visibility')
        print('8. Mean pressure')
        print('9. Wind direction')

        choice2 = int(input('Enter your choice: '))
        if choice2 == 1:
            if i == 0:
                data1name = data1name + ' Mean temperature (C)'
            elif i == 1:
                data2name = data2name + ' Mean temperature (C)'
            for j in range(7, 190, offset):
                data = str(sheet_ranges['B' + str(j)].value)
                if data == 'n/a':
                    print("WARNING: n/a value found. This may mess up your data.")
                    continue
                data = data.replace(' ', '.')
                data = float(data)
                print(data)
                if i == 0:
                    data1.append(data)
                    date1.append(sheet_ranges.cell(row=j, column=1).value)
                elif i == 1:
                    data2.append(data)
                    date2.append(sheet_ranges.cell(row=j, column=1).value)
        elif choice2 == 2:
            if i == 0:
                data1name = data1name + ' Total Rainfall (mm)'
            elif i == 1:
                data2name = data2name + ' Total Rainfall (mm)'
            for j in range(7, 190, offset):
                data = str(sheet_ranges['C' + str(j)].value)
                if data == 'n/a':
                    print("WARNING: n/a value found. This may mess up your data.")
                    continue
                data = data.replace(' ', '.')
                data = float(data)
                print(data)
                if i == 0:
                    data1.append(data)
                    date1.append(sheet_ranges.cell(row=j, column=1).value)
                elif i == 1:
                    data2.append(data)
                    date2.append(sheet_ranges.cell(row=j, column=1).value)
        elif choice2 == 3:
            if i == 0:
                data1name = data1name + ' Total Sunshine (hours)'
            elif i == 1:
                data2name = data2name + ' Total Sunshine (hours)'
            for j in range(7, 190, offset):
                data = str(sheet_ranges['D' + str(j)].value)
                if data == 'n/a':
                    print("WARNING: n/a value found. This may mess up your data.")
                    continue
                data = data.replace(' ', '.')
                data = float(data)
                print(data)
                if i == 0:
                    data1.append(data)
                    date1.append(sheet_ranges.cell(row=j, column=1).value)
                elif i == 1:
                    data2.append(data)
                    date2.append(sheet_ranges.cell(row=j, column=1).value)
        elif choice2 == 4:
            print("1. Wind speed")
            print("2. Wind speed (Beufort conversion)")
            print("3. Wind gust")
            choice3 = int(input('Enter your choice: '))
            if choice3 == 1:
                if i == 0:
                    data1name = data1name + ' Mean wind speed (kn)'
                elif i == 1:
                    data2name = data2name + ' Mean wind speed (kn)'
                for j in range(7, 190, offset):
                    data = str(sheet_ranges['E' + str(j)].value)
                    if data == 'n/a':
                        print("WARNING: n/a value found. This may mess up your data.")
                        continue
                    data = data.replace(' ', '.')
                    data = float(data)
                    print(data)
                    if i == 0:
                        data1.append(data)
                        date1.append(sheet_ranges.cell(row=j, column=1).value)
                    elif i == 1:
                        data2.append(data)
                        date2.append(sheet_ranges.cell(row=j, column=1).value)
            elif choice3 == 2:
                if i == 0:
                    data1name = data1name + ' Mean wind speed (Beaufort conversion)'
                elif i == 1:
                    data2name = data2name + ' Mean wind speed (Beaufort conversion)'
                for j in range(7, 190, offset):
                    data = str(sheet_ranges['F' + str(j)].value)
                    if data == 'n/a':
                        print("WARNING: n/a value found. This may mess up your data.")
                        continue
                    data = data.replace(' ', '.')
                    data = float(data)
                    print(data)
                    if i == 0:
                        data1.append(data)
                        date1.append(sheet_ranges.cell(row=j, column=1).value)
                    elif i == 1:
                        data2.append(data)
                        date2.append(sheet_ranges.cell(row=j, column=1).value)
            elif choice3 == 3:
                if i == 0:
                    data1name = data1name + ' Mean wind gust (kn)'
                elif i == 1:
                    data2name = data2name + ' Mean wind gust (kn)'
                for j in range(7, 190, offset):
                    data = str(sheet_ranges['G' + str(j)].value)
                    if data == 'n/a':
                        print("WARNING: n/a value found. This may mess up your data.")
                        continue
                    data = data.replace(' ', '.')
                    data = float(data)
                    print(data)
                    if i == 0:
                        data1.append(data)
                        date1.append(sheet_ranges.cell(row=j, column=1).value)
                    elif i == 1:
                        data2.append(data)
                        date2.append(sheet_ranges.cell(row=j, column=1).value)
        elif choice2 == 5:
            if i == 0:
                data1name = data1name + ' Relative humidity (%)'
            elif i == 1:
                data2name = data2name + ' Relative humidity (%)'
            for j in range(7, 190, offset):
                data = str(sheet_ranges['H' + str(j)].value)
                if data == 'n/a':
                    print("WARNING: n/a value found. This may mess up your data.")
                    continue
                data = data.replace(' ', '.')
                data = float(data)
                print(data)
                if i == 0:
                    data1.append(data)
                    date1.append(sheet_ranges.cell(row=j, column=1).value)
                elif i == 1:
                    data2.append(data)
                    date2.append(sheet_ranges.cell(row=j, column=1).value)
        elif choice2 == 6:
            if i == 0:
                data1name = data1name + ' Mean cloud cover (oktas)'
            elif i == 1:
                data2name = data2name + ' Mean cloud cover (oktas)'
            for j in range(7, 190, offset):
                data = str(sheet_ranges['I' + str(j)].value)
                if data == 'n/a':
                    print("WARNING: n/a value found. This may mess up your data.")
                    continue
                data = data.replace(' ', '.')
                data = float(data)
                print(data)
                if i == 0:
                    data1.append(data)
                    date1.append(sheet_ranges.cell(row=j, column=1).value)
                elif i == 1:
                    data2.append(data)
                    date2.append(sheet_ranges.cell(row=j, column=1).value)
        elif choice2 == 7:
            if i == 0:
                data1name = data1name + ' Mean visibility (dam)'
            elif i == 1:
                data2name = data2name + ' Mean visibility (dam)'
            for j in range(7, 190, offset):
                data = str(sheet_ranges['J' + str(j)].value)
                if data == 'n/a':
                    print("WARNING: n/a value found. This may mess up your data.")
                    continue
                data = data.replace(' ', '.')
                data = float(data)
                print(data)
                if i == 0:
                    data1.append(data)
                    date1.append(sheet_ranges.cell(row=j, column=1).value)
                elif i == 1:
                    data2.append(data)
                    date2.append(sheet_ranges.cell(row=j, column=1).value)
        elif choice2 == 8:
            if i == 0:
                data1name = data1name + ' Mean pressure (hPa)'
            elif i == 1:
                data2name = data2name + ' Mean pressure (hPa)'
            for j in range(7, 190, offset):
                data = str(sheet_ranges['K' + str(j)].value)
                if data == 'n/a':
                    print("WARNING: n/a value found. This may mess up your data.")
                    continue
                data = data.replace(' ', '.')
                data = float(data)
                print(data)
                if i == 0:
                    data1.append(data)
                    date1.append(sheet_ranges.cell(row=j, column=1).value)
                elif i == 1:
                    data2.append(data)
                    date2.append(sheet_ranges.cell(row=j, column=1).value)
        elif choice2 == 9:
            print("1. Wind direction (degrees)")
            print("2. Wind direction (compass)")
            print("3. Wind Gust direction (degrees)")
            print("4. Wind Gust direction (compass)")
            choice4 = int(input('Enter your choice: '))
            if choice4 == 1:
                if i == 0:
                    data1name = data1name + ' Mean wind direction (degrees)'
                elif i == 1:
                    data2name = data2name + ' Mean wind direction (degrees)'
                for j in range(7, 190, offset):

                    data = str(sheet_ranges['L' + str(j)].value)
                    if data == 'n/a':
                        print("WARNING: n/a value found. This may mess up your data.")
                        continue
                    data = data.replace(' ', '.')
                    data = float(data)
                    print(data)
                    if i == 0:
                        data1.append(data)
                        date1.append(sheet_ranges.cell(row=j, column=1).value)
                    elif i == 1:
                        data2.append(data)
                        date2.append(sheet_ranges.cell(row=j, column=1).value)
            elif choice4 == 2:
                if i == 0:
                    data1name = data1name + ' Mean wind direction (compass)'
                elif i == 1:
                    data2name = data2name + ' Mean wind direction (compass)'
                for j in range(7, 190, offset):
                    data = str(sheet_ranges['M' + str(j)].value)
                    if data == 'n/a':
                        print("WARNING: n/a value found. This may mess up your data.")
                        continue
                    data = data.replace(' ', '.')
                    data = float(data)
                    print(data)
                    if i == 0:
                        data1.append(data)
                    elif i == 1:
                        data2.append(data)
            elif choice4 == 3:
                if i == 0:
                    data1name = data1name + ' Mean wind gust direction (degrees)'
                elif i == 1:
                    data2name = data2name + ' Mean wind gust direction (degrees)'
                for j in range(7, 190, offset):
                    data = str(sheet_ranges['N' + str(j)].value)
                    if data == 'n/a':
                        print("WARNING: n/a value found. This may mess up your data.")
                        continue
                    data = data.replace(' ', '.')
                    data = float(data)
                    print(data)
                    if i == 0:
                        data1.append(data)
                        date1.append(sheet_ranges.cell(row=j, column=1).value)
                    elif i == 1:
                        data2.append(data)
                        date2.append(sheet_ranges.cell(row=j, column=1).value)
            elif choice4 == 4:
                if i == 0:
                    data1name = data1name + ' Mean wind gust direction (compass)'
                elif i == 1:
                    data2name = data2name + ' Mean wind gust direction (compass)'
                for j in range(7, 190, offset):
                    data = str(sheet_ranges['O' + str(j)].value)
                    if data == 'n/a':
                        print("WARNING: n/a value found. This may mess up your data.")
                        continue
                    data = data.replace(' ', '.')
                    data = float(data)
                    print(data)
                    if i == 0:
                        data1.append(data)
                        date1.append(sheet_ranges.cell(row=j, column=1).value)
                    elif i == 1:
                        data2.append(data)
                        date2.append(sheet_ranges.cell(row=j, column=1).value)
        # increment i
        if i == 0:
            i += 1
        elif i == 1:
            i += 1
    # debugging purposes

    if len(data1) != len(data2):
        print("WARNING: Data sets are not the same length. This will cause an error when you make a graph.")


    print('Data1: ' + str(data1))
    print('Data2: ' + str(data2))
    print('Data1name: ' + str(data1name))
    print('Data2name: ' + str(data2name))
    print('Date1: ' + str(date1))
    print('Date2: ' + str(date2))

    return data1, data2, data1name, data2name, date1, date2
